# 1. 로또 번호 크롤러
#  - 1회차 크롤링
#  - 다회차 크롤링
# 2. 엑셀에 데이터 저장
#  - 아무값이나 엑셀에 써보기
#  - 특정 회차 번호 저장해보기
#  - 다회차 저장해보기

import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup

# 1~856회차까지 로또 번호를 수집
# 엑셀 파일 저장
# 1 : ?, 2 : ?  <- 번호 카운팅

wb = Workbook()
ws1 = wb.active
ws2 = wb.create_sheet(title='second sheet')

# ws1.cell(row=1, column=1, value=1) # A1  ws1['A1']

# comprehension 내포식
numbers_count = {x:0 for x in range(1,46)}

for num in range(1,4):
    url ="https://search.daum.net/search?w=tot&rtmaxcoll=LOT&DA=LOT&q="+str(num)+"회차%20로또"
    req = requests.get(url)

    if req.status_code == requests.codes.ok:
        # print('접속 성공')
        html = BeautifulSoup(req.text, "html.parser")
        numbers = html.select('.lottonum .img_lotto')
        ws2.cell(row=num, column=1, value=f"{num}회차")

        for number in numbers[:6]:
            print(number.text, end=", ")
        print(numbers[-1].text)

        for index, number in enumerate(numbers):
            try:
                numbers_count[int(number.text)] += 1
                ws2.cell(row=num, column=index+2, value=number.text)
            except:
                pass
    else:
        print('접속 실패')

for number in numbers_count:
    print(number,":",numbers_count[number])

for row in range(1,46):
    ws1.cell(row=row, column=1, value=row)
    ws1.cell(row=row, column=2, value=numbers_count[row])

wb.save('test.xlsx')